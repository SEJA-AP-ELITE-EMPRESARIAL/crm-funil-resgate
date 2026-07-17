"""
Importa a planilha de INDICAÇÕES do APN (aba "Indicados") para o funil.

    python manage.py importar_indicados "caminho/arquivo.xlsx" [--aba Indicados]
        [--funil indicados_apn] [--dry-run] [--limpar]

Cada linha da aba = 1 indicado (lead). Mapeia:
  Nome/Empresa/WhatsApp Indicado  -> nome (Nome — Empresa), telefone
  Prioridade (P1..P5)             -> status + ordem (P1 no topo)
  Faixa de Faturamento            -> segmento
  Status (vazio)                  -> etapa "Priorizado"
  Indicador/Equipe/Obs.           -> notas (contexto preservado)
"""
import unicodedata

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.crm.models import Cliente, EtapaFunil, Funil

User = get_user_model()


def _norm(s):
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def _v(row, idx, col):
    i = idx.get(col)
    if i is None or i >= len(row):
        return ""
    val = row[i]
    return "" if val is None else str(val).strip()


class Command(BaseCommand):
    help = "Importa indicados do APN (aba 'Indicados') para o CRM."

    def add_arguments(self, parser):
        parser.add_argument("arquivo")
        parser.add_argument("--aba", default="Indicados")
        parser.add_argument("--funil", default="indicados_apn")
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--limpar", action="store_true",
                            help="Apaga os clientes existentes do funil antes de importar.")

    def handle(self, *args, **o):
        try:
            wb = load_workbook(o["arquivo"], read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f"Não foi possível abrir o arquivo: {e}")
        if o["aba"] not in wb.sheetnames:
            raise CommandError(f"Aba '{o['aba']}' não encontrada. Abas: {wb.sheetnames}")
        ws = wb[o["aba"]]

        try:
            funil = Funil.objects.get(slug=o["funil"])
        except Funil.DoesNotExist:
            raise CommandError(f"Funil '{o['funil']}' não existe.")

        rows = list(ws.iter_rows(values_only=True))
        hdr = rows[0]
        # Cabeçalhos podem repetir (ex.: "Prioridade" e "Faixa" têm colunas
        # auxiliares). Mantém a PRIMEIRA ocorrência = a coluna principal.
        idx = {}
        for i, h in enumerate(hdr):
            if h:
                k = _norm(h)
                idx.setdefault(k, i)

        # colunas esperadas (por nome normalizado)
        C = {k: _norm(k) for k in [
            "Nome Indicado", "Empresa Indicado", "WhatsApp Indicado",
            "Nome Indicador", "Empresa Indicador", "WhatsApp Indicador",
            "Equipe do Indicador", "Faixa de Faturamento (proxy)", "Prioridade",
            "Qtd. de Indicações do Indicador", "Status", "Força da Ponte",
            "Responsável", "Observações",
        ]}

        etapas = {_norm(v): v for v, _ in EtapaFunil.choices}
        etapas.update({_norm(l): v for v, l in EtapaFunil.choices})

        criados, ignorados, preview = [], 0, []
        for row in rows[1:]:
            if not any(c not in (None, "") for c in row):
                continue
            nome_ind = _v(row, idx, C["Nome Indicado"])
            empresa = _v(row, idx, C["Empresa Indicado"])
            if not nome_ind and not empresa:
                ignorados += 1
                continue

            nome = nome_ind
            if empresa and _norm(empresa) != _norm(nome_ind):
                nome = f"{nome_ind} — {empresa}" if nome_ind else empresa
            nome = nome[:200]

            prioridade = _v(row, idx, C["Prioridade"]).upper()
            ordem = 9
            if prioridade[:1] == "P" and prioridade[1:2].isdigit():
                ordem = int(prioridade[1:2])

            status_val = _v(row, idx, C["Status"])
            etapa = etapas.get(_norm(status_val), EtapaFunil.PRIORIZADO)

            qtd_txt = _v(row, idx, C["Qtd. de Indicações do Indicador"])
            try:
                qtd = int(float(qtd_txt)) if qtd_txt else None
            except ValueError:
                qtd = None

            c = Cliente(
                funil=funil,
                nome=nome,
                telefone=_v(row, idx, C["WhatsApp Indicado"])[:40],
                etapa=etapa,
                ordem=ordem,
                prioridade=prioridade[:10],
                faixa_faturamento=_v(row, idx, C["Faixa de Faturamento (proxy)"])[:60],
                indicador_nome=_v(row, idx, C["Nome Indicador"])[:120],
                indicador_empresa=_v(row, idx, C["Empresa Indicador"])[:160],
                indicador_whatsapp=_v(row, idx, C["WhatsApp Indicador"])[:40],
                indicador_equipe=_v(row, idx, C["Equipe do Indicador"])[:120],
                qtd_indicacoes=qtd,
                quem_fara_contato=_v(row, idx, C["Responsável"])[:120],
                notas=_v(row, idx, C["Observações"])[:2000],
            )
            criados.append(c)
            if len(preview) < 3:
                preview.append(c)

        self.stdout.write(f"Aba '{o['aba']}': {len(criados)} a importar, {ignorados} ignorados (sem nome/empresa).")
        self.stdout.write("--- prévia (3 primeiros) ---")
        for c in preview:
            self.stdout.write(f"  • nome='{c.nome}' | tel='{c.telefone}' | etapa={c.etapa} | "
                              f"prioridade={c.prioridade} | faixa='{c.faixa_faturamento}' | ordem={c.ordem}")
            self.stdout.write(f"    indicador='{c.indicador_nome}' ({c.indicador_empresa}) | "
                              f"equipe='{c.indicador_equipe}' | qtd={c.qtd_indicacoes} | notas='{c.notas[:60]}'")

        if o["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY-RUN: nada foi gravado."))
            return

        if o["limpar"]:
            n, _ = Cliente.objects.filter(funil=funil).delete()
            self.stdout.write(self.style.WARNING(f"{n} registro(s) do funil apagados antes de importar."))

        autor = User.objects.filter(is_superuser=True).first() or User.objects.first()
        for c in criados:
            c.criado_por = autor
            c.save()
        self.stdout.write(self.style.SUCCESS(f"{len(criados)} indicado(s) importado(s) no funil '{funil.nome}'."))
