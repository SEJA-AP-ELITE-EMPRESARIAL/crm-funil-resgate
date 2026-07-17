"""
Importação de clientes a partir de uma planilha Excel (.xlsx).

- Lê a primeira linha como cabeçalho e mapeia colunas (pt-BR, tolerante a
  acentos/maiúsculas) para os campos do Cliente.
- Resolve Funil (por nome ou slug) e Etapa (por rótulo ou slug).
- Valida linha a linha; devolve um resumo com criados + erros por linha.
- Também gera o modelo (.xlsx) de importação.
"""
import unicodedata
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook

from apps.crm.models import Cliente, EtapaFunil, Funil


def _norm(texto) -> str:
    """minúsculas, sem acento, sem espaços nas bordas."""
    s = str(texto or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s


# Cabeçalho normalizado -> campo do Cliente.
COLUNAS = {
    "funil": "funil",
    "nome": "nome", "empresa": "nome", "nome / empresa": "nome", "nome/empresa": "nome",
    "etapa": "etapa",
    "consultor": "quem_fara_contato", "consultor responsavel": "quem_fara_contato",
    "quem fara contato": "quem_fara_contato", "quem fara o contato": "quem_fara_contato",
    "responsavel": "responsavel",
    "telefone": "telefone", "fone": "telefone",
    "email": "email", "e-mail": "email",
    "cnpj": "cnpj",
    "segmento": "segmento",
    "canal": "canal",
    "status": "status",
    "municipio": "municipio", "cidade": "municipio",
    "estado": "estado", "uf": "estado",
    "pais": "pais",
    "produto": "produto_atual", "produto atual": "produto_atual",
    "consultor atual": "consultor_atual",
    "motivo": "motivo_distrato", "motivo distrato": "motivo_distrato",
    "motivo do distrato": "motivo_distrato",
    "notas": "notas", "observacoes": "notas", "obs": "notas",
    "valor": "valor_contrato", "valor contrato": "valor_contrato",
    "valor do contrato": "valor_contrato",
    "meses": "meses_contrato", "meses contrato": "meses_contrato",
    "duracao": "meses_contrato", "duracao do contrato": "meses_contrato",
    "data onboarding": "data_onboarding", "data offboarding": "data_offboarding",
    "qtd socios": "qtd_socios", "socios": "qtd_socios",
    "lt": "lt",
}

# Colunas do modelo, na ordem de exibição.
MODELO_COLUNAS = [
    "Funil", "Nome / Empresa", "Etapa", "Consultor", "Telefone", "Email", "CNPJ",
    "Segmento", "Municipio", "Estado", "Produto atual", "Motivo distrato",
    "Valor contrato", "Meses contrato", "Notas",
]

MODELO_EXEMPLO = [
    "Resgate", "Empresa Exemplo Ltda", "Priorizado", "Ana Souza", "(11) 99999-0000",
    "contato@exemplo.com", "12.345.678/0001-90", "Serviços", "São Paulo", "SP",
    "Plano Pro", "Preço", "12000,00", "12", "Cliente veio por indicação.",
]


def _parse_valor(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace("R$", "").strip()
    if "," in s:  # formato pt-BR
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError(f"valor inválido: {v!r}")


def _parse_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        raise ValueError(f"número inválido: {v!r}")


def _mapa_etapas():
    m = {}
    for value, label in EtapaFunil.choices:
        m[_norm(value)] = value
        m[_norm(label)] = value
    return m


def _mapa_funis():
    m = {}
    for f in Funil.objects.all():
        m[_norm(f.nome)] = f
        m[_norm(f.slug)] = f
    return m


def gerar_modelo_xlsx() -> bytes:
    """Planilha modelo com cabeçalho + 1 linha de exemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(MODELO_COLUNAS)
    ws.append(MODELO_EXEMPLO)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_clientes(arquivo, usuario=None) -> dict:
    """Lê o .xlsx e cria os clientes. Retorna {criados, total, erros:[...]}"""
    try:
        wb = load_workbook(arquivo, read_only=True, data_only=True)
    except Exception as exc:  # arquivo inválido
        raise ValueError(f"Não foi possível ler o arquivo: {exc}")

    ws = wb.active
    linhas = ws.iter_rows(values_only=True)

    try:
        cabecalho = next(linhas)
    except StopIteration:
        return {"criados": 0, "total": 0, "erros": [{"linha": 1, "erro": "Planilha vazia."}]}

    # Mapeia índice de coluna -> campo
    campos_por_indice = {}
    for i, titulo in enumerate(cabecalho):
        campo = COLUNAS.get(_norm(titulo))
        if campo:
            campos_por_indice[i] = campo

    if "nome" not in campos_por_indice.values():
        raise ValueError("A planilha precisa ter uma coluna 'Nome / Empresa'.")

    etapas = _mapa_etapas()
    funis = _mapa_funis()

    criados = 0
    erros = []
    a_criar = []

    for n, linha in enumerate(linhas, start=2):  # linha 2 = primeira de dados
        if linha is None or all(c is None or str(c).strip() == "" for c in linha):
            continue  # linha em branco
        dados = {}
        erro_linha = None
        for i, campo in campos_por_indice.items():
            valor = linha[i] if i < len(linha) else None
            if valor is None or str(valor).strip() == "":
                continue
            try:
                if campo == "funil":
                    f = funis.get(_norm(valor))
                    if not f:
                        raise ValueError(f"funil '{valor}' não encontrado")
                    dados["funil"] = f
                elif campo == "etapa":
                    e = etapas.get(_norm(valor))
                    if not e:
                        raise ValueError(f"etapa '{valor}' inválida")
                    dados["etapa"] = e
                elif campo == "valor_contrato":
                    dados["valor_contrato"] = _parse_valor(valor)
                elif campo in ("meses_contrato", "qtd_socios"):
                    dados[campo] = _parse_int(valor)
                else:
                    dados[campo] = str(valor).strip()
            except ValueError as e:
                erro_linha = f"coluna '{cabecalho[i]}': {e}"
                break

        if erro_linha:
            erros.append({"linha": n, "erro": erro_linha})
            continue
        if not dados.get("nome"):
            erros.append({"linha": n, "erro": "nome vazio"})
            continue

        cliente = Cliente(**dados)
        if usuario is not None and getattr(usuario, "is_authenticated", False):
            cliente.criado_por = usuario
        a_criar.append(cliente)

    with transaction.atomic():
        for c in a_criar:
            c.save()
            criados += 1

    return {"criados": criados, "total": criados + len(erros), "erros": erros}
